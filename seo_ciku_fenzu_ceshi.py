#coding:utf-8
# 创建时间：2017年10月27日13:55:54
# 脚本功能：测试分词分组
# 2017年11月03日16:45:00
# 完成词库分组，调用key_list.xlsx 文件中的原始关键词及指数，进行tf-idf分组
# 2017年11月03日17:25:34 测试最终版功能无误，在大量数据处理上速度较慢，50万数据分组20分钟。
# 依赖库：jieba（分词） openpyxl（excel操作） numpy pandas（分组）
# 2017年12月06日11:42:56 v2.0
# 测试改写为字典格式，提高分组效率

import jieba,sys, datetime
import jieba.analyse
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import Border, Side, Font
from openpyxl.cell import get_column_letter
from optparse import OptionParser
from openpyxl.reader.excel import load_workbook
import pandas as pd
import numpy as np


reload(sys)
sys.setdefaultencoding("utf-8")


print '———————— 1. 开始运行 ————————'
begin = datetime.datetime.now()
print '开始时间：{}'.format(begin)
# text = open(r"ceshi.txt","r").read()
key_list = [] #原始关键词列表
zhishu_list = [] #原始关键词指数

d_key_index = {}

print '———————— 2.读取词库 ————————'

wb=load_workbook('key_list.xlsx') #从xlsx文件中获取词库数据

sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
max_row = ws.max_row # 读取表格所有行数

for row in range(2,max_row+1):
    key_list.append(str(ws.cell(row=row, column=1).value).replace('+', ' '))
    index = str(ws.cell(row=row, column=2).value)
    if index == 0:
        index = 1
    elif index == '未收录':
        index = 0
    zhishu_list.append(str(index))
    d_key_index[str(ws.cell(row=row, column=1).value).replace('+', ' ')] = int(index)
    # 从excel中拿到的数据，保存到d_key_index字典中


tfidf=jieba.analyse.extract_tags("\n".join(key_list),topK=100,withWeight=True) #通过TF-IDF算出30个高频分词

print '———————— 3.分词完毕 ————————'

stopkeyword = ['作文', '留学'] # 分组过滤词list

print '———————— 4.开始分组 ————————'

#对关键词进行分组
keywords = []
key_fenzu = []
key_zhishu = []

while len(d_key_index) > 0:
    k = d_key_index.popitem()
    key = k[0]
    index = k[1]
    # 通过 dict.popitem() 方法拿到字典中的数据
    for fenzu in tfidf:
        if fenzu[0] in stopkeyword:
            pass #如果分组属于过滤词就pass
        elif fenzu[0] in key:
            # print "key："+key+','+"分组："+fenzu[0]
            keywords.append(key)
            key_fenzu.append(fenzu[0])
            key_zhishu.append(int(index))
            # print zhishu_list[key_list.txt.index(key)]
            break

# print keywords[3]
# print key_zhishu[3]
# print key_fenzu[3]

print '———————— 5.计算总数据 ————————'

# 通过pandas把指数最大的30个分组导出到output.xlsx文件中
data = pd.DataFrame({"关键词":keywords,"分组":key_fenzu,"指数":key_zhishu})
df = data.groupby(by=['分组'])['指数'].agg({
        '指数': np.sum,
        '词量': np.size,
        '平均': np.mean
    })
df_sort = df.sort_values(['指数'], ascending=False)  # 排序,以总得分降序
# df_sort.to_csv("./test_fenzu_zuming.csv")
writer = pd.ExcelWriter('output.xlsx')
df_sort.to_excel(writer,'Sheet1')
writer.save()

print '———————— 6.总数据导出 output.xlsx ————————'

# print data.pivot_table(index='分组', values=['指数'], aggfunc=sum).sort_values(['指数'], ascending=False) #数据透视表 pivot_table

print '———————— 7.开始分组归类 ————————'

# 读取output.xlsx文件，获得最大指数分组的顺序，与关键词和指数
wb=load_workbook("./output.xlsx")
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
#读取B1单元格中的内容
# print ws.cell("A2").value
row = ws.max_row #获取表格行数
out_key =[]
out_index = []
for i in range(2,int(row)+1):
    key = ws.cell(row=i, column=1).value
    index = ws.cell(row=i, column=3).value
    # print key,index
    out_key.append(key)
    out_index.append(index)

# print ws.cell(row=2, column=1).value

# get_fenzu_key 方法，获取所有要保存的数据，以字典形式返回
def get_fenzu_key(fenzu):
    save_dict = {}
    for index,val in enumerate(key_fenzu):
        if val == fenzu:
            # print keywords[index],key_zhishu[index],fenzu
            save_dict.setdefault(keywords[index],key_zhishu[index])
    return save_dict

# 获得总指数方法
def alindex():
    i = 0
    for index in out_index:
        i = i + index
    return i

# 获取单个分组方法
def get_zhanbi(index):
    all_index = alindex()
    zhanbi = float(index)/float(all_index)
    res = format(zhanbi, '.0%')
    return res


# save_list = get_fenzu_key(out_key[0])
# print save_list

# save_list = sorted(save_list.items(), lambda x, y: cmp(x[1], y[1]), reverse=True)
# print save_list

# for key,value in save_list:
#    print key, value # print key,dict[key]

print '———————— 8.分组导出 ————————'

# 新建一个workbook
wb = Workbook()
# 新建一个excelWriter
ew = ExcelWriter(workbook=wb)
# 设置文件输出路径与名称
dest_filename = r'ciku_list.xlsx'
# 第一个sheet是ws
ws = wb.worksheets[0]
# 设置ws的名称
ws.title = "keylist"

''' 测试的时候写的单个分组内容 '''
# ws.cell(row=1, column=1).value = out_key[0]+'('+get_zhanbi(out_index[0])+')'
# ws.cell(row=1, column=2).value = out_index[0]
# xunhuan = 1
# for key,value in save_list:
#     xunhuan = xunhuan+1
#     ws.cell(row=xunhuan, column=1).value = key
#     ws.cell(row=xunhuan, column=2).value = value

# 设置表格格式
font = Font(size=12, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
border = Border(top=Side(style='medium', color='FF000000'), bottom=Side(style='medium', color='FF000000'),
                diagonal=Side(style='medium', color='FF000000'), diagonal_direction=0,
                outline=Side(style='medium', color='FF000000'), vertical=Side(style='medium', color='FF000000'),
                horizontal=Side(style='medium', color='FF000000'))
border_right = Border(right=Side(border_style="thin", color='FF000000'),diagonal=Side(style='medium', color='FF000000'), diagonal_direction=0,
                outline=Side(style='medium', color='FF000000'), vertical=Side(style='medium', color='FF000000'),
                horizontal=Side(style='medium', color='FF000000'))
fist_border_right = Border(right=Side(border_style="thin", color='FF000000'),top=Side(style='medium', color='FF000000'), bottom=Side(style='medium', color='FF000000'),diagonal=Side(style='medium', color='FF000000'), diagonal_direction=0,
                outline=Side(style='medium', color='FF000000'), vertical=Side(style='medium', color='FF000000'),
                horizontal=Side(style='medium', color='FF000000'))

# 两次循环输出，第一次循环输出第一行数据，总的分组、分组占比、分组总指数
# 第二次循环输出，对应的分词关键词、关键词所属指数
col_key = 1
col_ind = 2
for i in range(0,len(out_key)):
    # print out_key[i],i
    if i ==0:
        col_key = 1
        col_ind = 2
    else:
        col_key = col_key + 2
        col_ind = col_ind + 2
    ws.cell(row=1, column=col_key).value = out_key[i] + '（' + get_zhanbi(out_index[i]) + '）'
    ws.cell(row=1, column=col_ind).value = out_index[i]
    ws.cell(row=1, column=col_key).font = font
    ws.cell(row=1, column=col_key).border = border
    ws.cell(row=1, column=col_ind).font = font
    ws.cell(row=1, column=col_ind).border = fist_border_right

    save_list = get_fenzu_key(out_key[i])
    save_list = sorted(save_list.items(), lambda x, y: cmp(x[1], y[1]), reverse=True)
    xunhuan = 1
    for key,value in save_list:
        xunhuan = xunhuan+1
        ws.cell(row=xunhuan, column=col_key).value = key
        ws.cell(row=xunhuan, column=col_ind).value = value
        ws.cell(row=xunhuan, column=col_ind).border = border_right

# 最后保存文件


ew.save(filename=dest_filename)
print '———————— 9.分组导出完毕 ————————'
print '———————— 10.程序结束 ————————'
end = datetime.datetime.now()
print '结束时间：{}'.format(end)
print '程序用时：{}秒'.format(str((end - begin).seconds))
